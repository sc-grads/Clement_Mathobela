#################################
## Working with Excel Files in Python
## The OpenPyXL Module
## Sheets Operations
#################################

## Importing the module
import openpyxl

## Loading an excel file. store.xlsx is in the same directory with the python script
wb = openpyxl.load_workbook('store.xlsx')

## Printing all sheets by name
print(wb.sheetnames)

## Getting a sheet by name
sheet = wb['Products']

## Printing all sheet's methods
print(dir(sheet))

## Changing the sheet title
# sheet.title = 'Products for sale'


## Printing sheet properties
print(sheet.sheet_format)
print(sheet.sheet_properties)

## Create a new sheet on first position in the Workbook (by default will be the last)
wb.create_sheet('Turnover1', 0)

## Getting a sheet by name
sheet1 = wb['Turnover1']

## Removing a sheet by name
wb.remove(sheet1)

## Copying sheets in the same Workbook
source = wb['Turnover']
destination = wb.copy_worksheet(source)
# print(destination.title)

## Saving the changes
wb.save('store.xlsx')