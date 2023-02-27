#################################
## Working with Excel Files in Python
## The OpenPyXL Module
## Reading Excel Files
#################################

## Importing the module
import openpyxl

## Loading an excel file. store.xlsx is in the same directory with the python script
## stored.xlsx is attached to this lecture
wb = openpyxl.load_workbook('store.xlsx')  # wb is a WorkBook object

## Loading an excel file. If there are formulas, values (results of formulas) are read and not the formulas
# wb = openpyxl.load_workbook('store.xlsx', data_only=True)

## Printing all sheet names
print(wb.sheetnames)  # sheetnames is a list

## Iterating of the sheets in the workbook and printing their title
for sheet in wb:
    print(sheet.title)

## Getting a sheet by name. All subsequent operations will be done on this sheet
sheet = wb['Products']

## Getting the active sheet. By default it's the last opened sheet
sheet = wb.active

## Getting cells
b2_cell = sheet['B2']
c2_cell = sheet['c2']

## Getting values of cells
print(b2_cell.value, c2_cell.value)

## Another way to get the values of cells
print(sheet.cell(row=4, column=2).value)
print(c2_cell.row, c2_cell.column)

## Getting the data type
print(sheet['A5'].data_type)
print(sheet['B5'].data_type)

## Getting the encoding format of cell
print(sheet['A5'].encoding)

print(sheet['D4'].parent)

## Printing all cell methods
print(dir(b2_cell))

## Getting a cell range
cell_range = sheet['B2:C11']

## Iterating over the cell range
for product, price in cell_range:
    print(f'Product: {product.value}    Price:{price.value}')

## Getting the sheet dimensions
print(f'Sheet Dimentions: {sheet.dimensions}')
print(sheet.max_row, sheet.max_column)

## Iterating and getting all data in a sheet
for a, b, c, d, e in sheet[sheet.dimensions]:
    print(a.value, b.value, c.value, d.value, e.value)

for row in sheet.rows:
    for cell in row:
        print(f'{cell.value} ', end='')
    print('\n')

## Another way to get all data in sheet. every row is a tuple
for row in sheet.values:
    print(row)